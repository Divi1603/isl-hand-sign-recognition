"""
ISL GESTURE RECOGNITION — STREAMLIT APP (FIXED)
KEY FIX: Feature extraction now matches training (166 features, two-hand)
Run: streamlit run 3_app.py
"""

import streamlit as st
import cv2
import mediapipe as mp
import numpy as np
import pickle
import pandas as pd
import json, time, os
from datetime import datetime
from collections import deque, Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

st.set_page_config(page_title="ISL Gesture Recognition", page_icon="🤟", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body,
[data-testid="stAppViewContainer"],
[data-testid="stApp"],
.main,
.block-container,
section.main > div {
    background-color: #ffffff !important;
    color: #111827 !important;
    font-family: 'Inter', sans-serif !important;
}

[data-testid="stHeader"] {
    background-color: #ffffff !important;
    border-bottom: 1px solid #e5e7eb !important;
}

[data-testid="stSidebar"],
[data-testid="stSidebar"] > div {
    background-color: #f9fafb !important;
    border-right: 1px solid #e5e7eb !important;
}
[data-testid="stSidebar"] * { color: #111827 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span { color: #111827 !important; }
[data-testid="stSidebar"] [data-testid="stTickBarMin"],
[data-testid="stSidebar"] [data-testid="stTickBarMax"] { color: #6b7280 !important; }

h1, h2, h3, h4, h5, h6 { color: #111827 !important; font-family: 'Inter', sans-serif !important; }
p, span, label, div, li { color: #374151 !important; }
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li,
[data-testid="stMarkdownContainer"] strong { color: #111827 !important; }
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4 { color: #111827 !important; }

[data-testid="stMetric"] {
    background: #f0f4ff !important;
    border: 1px solid #c7d2fe !important;
    border-radius: 12px !important;
    padding: 14px 18px !important;
}
[data-testid="stMetricLabel"] > div {
    color: #6b7280 !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
[data-testid="stMetricValue"] > div {
    color: #111827 !important;
    font-size: 1.6rem !important;
    font-weight: 700 !important;
}

.big-gesture {
    font-size: 2.8rem;
    font-weight: 800;
    color: #1d4ed8 !important;
    text-align: center;
    padding: 18px 14px;
    border: 2.5px solid #3b82f6;
    border-radius: 14px;
    background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
    margin-bottom: 10px;
    letter-spacing: -0.5px;
    box-shadow: 0 2px 12px rgba(59,130,246,0.15);
}
.no-gesture {
    background: #f9fafb;
    border: 1.5px dashed #d1d5db;
    border-radius: 12px;
    padding: 18px 14px;
    text-align: center;
    color: #9ca3af !important;
    font-size: 1rem;
}

.sentence-box {
    background: #f9fafb;
    border-radius: 12px;
    padding: 16px 18px;
    font-size: 1.2rem;
    font-weight: 500;
    color: #111827 !important;
    min-height: 62px;
    border: 1.5px solid #e5e7eb;
    line-height: 1.6;
}
.sentence-placeholder {
    color: #9ca3af !important;
    font-style: italic;
}

.cam-placeholder {
    height: 400px;
    display: flex;
    align-items: center;
    justify-content: center;
    border: 2px dashed #d1d5db;
    border-radius: 14px;
    color: #9ca3af;
    font-size: 1.05rem;
    background: #fafafa;
    flex-direction: column;
    gap: 10px;
}

.stButton > button {
    background: #2563eb !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 9px !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    padding: 8px 16px !important;
    transition: background 0.2s ease !important;
    box-shadow: 0 1px 4px rgba(37,99,235,0.25) !important;
}
.stButton > button:hover {
    background: #1d4ed8 !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.35) !important;
}

[data-testid="stTabs"] [role="tablist"] { border-bottom: 2px solid #e5e7eb !important; }
button[data-baseweb="tab"] { color: #6b7280 !important; font-weight: 500 !important; font-size: 0.9rem !important; }
button[data-baseweb="tab"][aria-selected="true"] { color: #2563eb !important; border-bottom: 2.5px solid #2563eb !important; font-weight: 700 !important; }

[data-testid="stAlert"] { border-radius: 10px !important; }
[data-testid="stAlert"] p { color: inherit !important; }

[data-testid="stDataFrame"] * { color: #111827 !important; background-color: #ffffff !important; }
[data-testid="stDataFrame"] th { background-color: #f3f4f6 !important; color: #374151 !important; font-weight: 600 !important; }

[data-testid="stProgress"] > div > div { background-color: #2563eb !important; }
[data-testid="stProgress"] p { color: #374151 !important; font-weight: 500 !important; }

hr { border-color: #e5e7eb !important; margin: 1rem 0 !important; }

[data-testid="stCheckbox"] span,
[data-testid="stCheckbox"] p { color: #111827 !important; font-weight: 500 !important; }
[data-testid="stSlider"] p,
[data-testid="stSlider"] span,
[data-testid="stSlider"] label { color: #111827 !important; }

[data-testid="stExpander"] summary p { color: #111827 !important; font-weight: 500 !important; }
[data-testid="stExpander"] { border: 1px solid #e5e7eb !important; border-radius: 10px !important; background: #f9fafb !important; }

[data-testid="stDownloadButton"] > button {
    background: #ffffff !important;
    color: #2563eb !important;
    border: 2px solid #2563eb !important;
    border-radius: 9px !important;
    font-weight: 600 !important;
}
[data-testid="stDownloadButton"] > button:hover { background: #eff6ff !important; }

.section-label {
    font-size: 0.78rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #6b7280 !important;
    margin-bottom: 6px;
    margin-top: 16px;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SHARED CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
CONNECTIONS = [
    (0,1),(1,2),(2,3),(3,4),(0,5),(5,6),(6,7),(7,8),
    (0,9),(9,10),(10,11),(11,12),(0,13),(13,14),(14,15),
    (15,16),(0,17),(17,18),(18,19),(19,20)
]
EXPECTED_FEATURES = 166

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════
def normalize_landmarks(lm_list):
    arr  = np.array([[l.x, l.y, l.z] for l in lm_list])
    arr -= arr[0].copy()
    scale = np.max(np.linalg.norm(arr, axis=1)) + 1e-6
    return arr / scale

def extract_single_hand(lm_list):
    arr   = normalize_landmarks(lm_list)
    dists = [float(np.linalg.norm(arr[a] - arr[b])) for a, b in CONNECTIONS]
    return arr[:,0].tolist() + arr[:,1].tolist() + arr[:,2].tolist() + dists

def get_empty_hand():
    return [0.0] * 83

def extract_features_two_hands(multi_hand_landmarks, multi_handedness):
    left_feats  = None
    right_feats = None
    if multi_hand_landmarks and multi_handedness:
        for lm, handedness in zip(multi_hand_landmarks, multi_handedness):
            label = handedness.classification[0].label
            feats = extract_single_hand(lm.landmark)
            if label == "Left":
                left_feats  = feats
            else:
                right_feats = feats
    left_feats  = left_feats  if left_feats  is not None else get_empty_hand()
    right_feats = right_feats if right_feats is not None else get_empty_hand()
    return np.array(left_feats + right_feats).reshape(1, -1)

# ── Model loader ──────────────────────────────────────────────────────────────
@st.cache_resource
def load_model():
    with open("isl_model.pkl",    "rb") as f: model  = pickle.load(f)
    with open("label_encoder.pkl","rb") as f: le     = pickle.load(f)
    with open("scaler.pkl",       "rb") as f: scaler = pickle.load(f)
    report = {}
    if os.path.exists("training_report.json"):
        with open("training_report.json") as f: report = json.load(f)
    n_model_features = model.estimators_[0].n_features_in_ if hasattr(model, 'estimators_') else None
    return model, le, scaler, report, n_model_features

MODEL_LOADED = False
model_err    = ""
try:
    model, le, scaler, report, n_model_features = load_model()
    MODEL_LOADED = True
    if n_model_features and n_model_features != EXPECTED_FEATURES:
        st.error(f"⚠️ Model expects {n_model_features} features but app uses {EXPECTED_FEATURES}. "
                 "Retrain with train_model.py!")
except Exception as e:
    model_err = str(e)

mp_hands    = mp.solutions.hands
mp_drawing  = mp.solutions.drawing_utils
mp_draw_sty = mp.solutions.drawing_styles

# ── Excel logging ─────────────────────────────────────────────────────────────
EXCEL_FILE = "isl_detections.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Detections"
        ws.append(["#","Timestamp","Gesture","Confidence (%)","Session ID"])
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for c in ws[1]:
            c.fill = fill; c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        for col, w in zip("ABCDE", [6,22,18,18,22]):
            ws.column_dimensions[col].width = w
        wb.save(EXCEL_FILE)

def log_excel(gesture, conf, session_id):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE); ws = wb["Detections"]
        n  = ws.max_row
        ws.append([n, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   gesture, f"{conf*100:.1f}", session_id])
        if n % 2 == 0:
            fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for c in ws[ws.max_row]: c.fill = fill
        wb.save(EXCEL_FILE)
    except Exception as ex:
        print(f"Excel log error: {ex}")

init_excel()

# ── Session state ─────────────────────────────────────────────────────────────
defaults = {
    "camera_on":  False,
    "sentence":   [],
    "last_added": "",
    "last_time":  0,
    "log":        [],
    "total":      0,
    "session_id": datetime.now().strftime("SES-%H%M%S"),
    "buf":        deque(maxlen=15),
    "fps_t":      deque(maxlen=30),
}
for k, v in defaults.items():
    if k not in st.session_state: st.session_state[k] = v

# ── Sidebar — settings only, no model info ────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    conf_thresh   = st.slider("Min Confidence",   0.30, 0.95, 0.45, 0.05)
    stable_frames = st.slider("Stability frames", 3,    15,   5)
    cooldown      = st.slider("Cooldown (s)",     0.5,  5.0,  1.5, 0.5)
    show_lm       = st.checkbox("Show landmarks",             True)
    show_fps      = st.checkbox("Show FPS",                   True)
    show_debug    = st.checkbox("Show debug overlay (top-3)", True)
    st.divider()

    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, "rb") as f:
            st.download_button(
                "⬇️ Download Excel Log", f, "isl_detections.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Main Header ───────────────────────────────────────────────────────────────
st.markdown("""
<div style="padding: 8px 0 16px 0;">
  <div style="font-size:2rem; font-weight:800; color:#111827 !important; letter-spacing:-0.5px; line-height:1.2;">
    🤟 Indian Sign Language — Real-Time Recognition
  </div>
  <div style="font-size:0.95rem; color:#6b7280; margin-top:6px;">
    <strong style="color:#374151;">Final Year Project</strong> &nbsp;|&nbsp;
    MediaPipe + Ensemble ML (two-hand, 166 features)
  </div>
</div>
""", unsafe_allow_html=True)
st.divider()

if not MODEL_LOADED:
    st.error(f"Cannot load model: {model_err}")
    st.info("Run `python train_model.py` first.")
    st.stop()

# Stats row
c1, c2, c3, c4 = st.columns(4)
stat_total = c1.empty()
stat_sess  = c2.empty()
stat_last  = c3.empty()
stat_fps   = c4.empty()

# Main columns
col_cam, col_info = st.columns([3, 2])

with col_cam:
    cam_ph = st.empty()
    b1, b2, b3 = st.columns(3)
    start_btn = b1.button("▶ Start Camera", type="primary", use_container_width=True)
    stop_btn  = b2.button("⏹ Stop Camera",                  use_container_width=True)
    clear_btn = b3.button("🗑 Clear Sentence",               use_container_width=True)

with col_info:
    st.markdown('<div class="section-label">🎯 Detected Gesture</div>', unsafe_allow_html=True)
    gest_ph  = st.empty()
    conf_ph  = st.empty()
    st.markdown('<div class="section-label">💬 Sentence</div>', unsafe_allow_html=True)
    sent_ph  = st.empty()
    undo_btn = st.button("↩ Undo last word")
    st.markdown('<div class="section-label">📋 Recent Detections</div>', unsafe_allow_html=True)
    table_ph = st.empty()

if start_btn: st.session_state.camera_on = True
if stop_btn:  st.session_state.camera_on = False
if clear_btn:
    st.session_state.sentence   = []
    st.session_state.last_added = ""
if undo_btn and st.session_state.sentence:
    st.session_state.sentence.pop()

# ── Render helper ─────────────────────────────────────────────────────────────
def render_ui(gesture="", conf=0.0, fps=0):
    stat_total.metric("Total",   st.session_state.total)
    stat_sess.metric("Session",  st.session_state.session_id)
    stat_last.metric("Last",     st.session_state.last_added or "—")
    stat_fps.metric("FPS",       f"{fps:.0f}")

    if gesture:
        gest_ph.markdown(
            f'<div class="big-gesture">{gesture}</div>', unsafe_allow_html=True)
        conf_ph.progress(int(conf * 100), text=f"Confidence: {conf*100:.0f}%")
    else:
        gest_ph.markdown(
            '<div class="no-gesture">No gesture detected</div>', unsafe_allow_html=True)
        conf_ph.empty()

    if st.session_state.sentence:
        words = " ".join(st.session_state.sentence)
        sent_ph.markdown(
            f'<div class="sentence-box">{words}</div>', unsafe_allow_html=True)
    else:
        sent_ph.markdown(
            '<div class="sentence-box sentence-placeholder">Start signing…</div>',
            unsafe_allow_html=True)

    if st.session_state.log:
        table_ph.dataframe(
            pd.DataFrame(st.session_state.log[-15:][::-1]),
            use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# CAMERA LOOP
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.camera_on:
    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT,  720)

    with mp_hands.Hands(
            static_image_mode=False,
            max_num_hands=2,
            min_detection_confidence=0.5,
            min_tracking_confidence=0.5) as hands_det:

        while st.session_state.camera_on:
            ret, frame = cap.read()
            if not ret: break

            frame = cv2.flip(frame, 1)
            now   = time.time()
            st.session_state.fps_t.append(now)
            fps = (len(st.session_state.fps_t) /
                   (st.session_state.fps_t[-1] - st.session_state.fps_t[0] + 1e-6))

            rgb    = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            result = hands_det.process(rgb)

            cur_gesture = ""
            cur_conf    = 0.0

            if result.multi_hand_landmarks:
                if show_lm:
                    for hand_lm in result.multi_hand_landmarks:
                        mp_drawing.draw_landmarks(
                            frame, hand_lm, mp_hands.HAND_CONNECTIONS,
                            mp_draw_sty.get_default_hand_landmarks_style(),
                            mp_draw_sty.get_default_hand_connections_style())

                try:
                    feats = extract_features_two_hands(
                        result.multi_hand_landmarks,
                        result.multi_handedness if result.multi_handedness else [])

                    if feats.shape[1] != EXPECTED_FEATURES:
                        cv2.putText(frame,
                            f"Feature mismatch: {feats.shape[1]} != {EXPECTED_FEATURES}",
                            (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,220), 2)
                    else:
                        fscale = scaler.transform(feats)
                        probs  = model.predict_proba(fscale)[0]
                        top_i  = int(np.argmax(probs))
                        top_c  = float(probs[top_i])
                        name   = le.classes_[top_i]

                        print(f">> {name} ({top_c*100:.1f}%)  thresh={conf_thresh*100:.0f}%")

                        if show_debug:
                            top3 = np.argsort(probs)[::-1][:3]
                            for rank, idx in enumerate(top3):
                                col  = (30, 180, 30) if rank == 0 else (120, 120, 120)
                                txt  = f"{le.classes_[idx]}: {probs[idx]*100:.0f}%"
                                cv2.putText(frame, txt, (10, 55 + rank*35),
                                            cv2.FONT_HERSHEY_DUPLEX, 0.85, col, 2)

                        h, w   = frame.shape[:2]
                        ph     = result.multi_hand_landmarks[0]
                        xs     = [l.x * w for l in ph.landmark]
                        ys     = [l.y * h for l in ph.landmark]
                        x1, y1 = max(0, int(min(xs))-20), max(0, int(min(ys))-20)
                        x2, y2 = min(w, int(max(xs))+20), min(h, int(max(ys))+20)

                        if top_c >= conf_thresh:
                            cur_gesture = name
                            cur_conf    = top_c
                            cv2.rectangle(frame, (x1,y1), (x2,y2), (37,99,235), 2)
                            cv2.putText(frame, f"{name}  {top_c*100:.0f}%",
                                        (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9,
                                        (37,99,235), 2)

                            st.session_state.buf.append(name)
                            most, cnt = Counter(st.session_state.buf).most_common(1)[0]
                            if cnt >= stable_frames:
                                t = time.time()
                                if (name != st.session_state.last_added or
                                        t - st.session_state.last_time > cooldown):
                                    st.session_state.sentence.append(name)
                                    st.session_state.last_added = name
                                    st.session_state.last_time  = t
                                    st.session_state.total     += 1
                                    log_excel(name, top_c, st.session_state.session_id)
                                    st.session_state.log.append({
                                        "Time":    datetime.now().strftime("%H:%M:%S"),
                                        "Gesture": name,
                                        "Conf%":   f"{top_c*100:.1f}"
                                    })
                        else:
                            cv2.rectangle(frame, (x1,y1), (x2,y2), (200,60,60), 2)
                            cv2.putText(frame, f"Low conf: {top_c*100:.0f}%",
                                        (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.75,
                                        (200,60,60), 2)
                            st.session_state.buf.clear()

                except Exception as ex:
                    print(f"Prediction error: {ex}")
                    cv2.putText(frame, f"Pred error: {ex}",
                                (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,200), 1)
            else:
                st.session_state.buf.clear()
                cv2.putText(frame, "No hands detected",
                            (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.85, (80,80,200), 2)

            h_count = len(result.multi_hand_landmarks) if result.multi_hand_landmarks else 0
            hc_col  = (20,160,80) if h_count else (80,80,200)
            cv2.putText(frame, f"Hands: {h_count}",
                        (frame.shape[1]-150, 35),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.85, hc_col, 2)

            if show_fps:
                cv2.putText(frame, f"FPS: {fps:.0f}",
                            (frame.shape[1]-150, 70),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.85, (0,140,180), 2)

            cam_ph.image(frame, channels="BGR", use_column_width=True)
            render_ui(cur_gesture, cur_conf, fps)

    cap.release()

else:
    cam_ph.markdown("""
    <div class="cam-placeholder">
        <span style="font-size:2.5rem;">📷</span>
        <span style="color:#6b7280; font-size:1rem;">
            Press <strong style="color:#2563eb;">▶ Start Camera</strong> to begin
        </span>
    </div>""", unsafe_allow_html=True)
    render_ui()

# ── Analytics tabs ────────────────────────────────────────────────────────────
st.divider()
tab1, tab2, tab3 = st.tabs(["📊 Analytics", "📁 Session Log", "ℹ️ How to Use"])

with tab1:
    c1, c2 = st.columns(2)
    with c1:
        if os.path.exists("confusion_matrix.png"):
            st.image("confusion_matrix.png", use_column_width=True)
        else:
            st.info("Train the model to see the confusion matrix.")
    with c2:
        if st.session_state.log:
            df_l = pd.DataFrame(st.session_state.log)
            st.bar_chart(df_l["Gesture"].value_counts())
        else:
            st.info("No detections in this session yet.")

with tab2:
    if os.path.exists(EXCEL_FILE):
        st.dataframe(pd.read_excel(EXCEL_FILE), use_container_width=True)
    else:
        st.info("No detections logged yet.")

with tab3:
    st.markdown("""
### Tips for best detection

- **Good lighting** — face a window or lamp; avoid dark rooms
- **Hold gesture steady** — keep hand still for 1–2 seconds before moving on
- **Full hand in frame** — don't let fingers go off the edge
- **Face palm toward camera** — not sideways
- **Lower confidence slider** if gestures are not triggering (try 0.40)
- **Blue box** = gesture accepted | **Red box** = confidence too low
- **Debug overlay** (top-3 predictions) shows on camera — see what the model is thinking
- If **no hands detected**, ensure your hand is fully visible and well-lit

### Feature note
This model uses **166 features** (both hands combined). Even single-hand gestures work because
the missing hand is automatically zero-padded — just like during training.
""")